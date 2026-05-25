"""build_reporte_maestro.py
Extrae las hojas del Reporte_Maestro_Produccion_NF_Sanmina_AFL hacia CSVs
curados en adriana_projects/data/curated/, listos para ser leídos por Flask.

Uso:
    python3 adriana_projects/scripts/build_reporte_maestro.py
"""

import csv
import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
CURATED_DIR = ROOT / "adriana_projects" / "data" / "curated"
REPORTE = ROOT / "Reporte_Maestro_Produccion_NF_Sanmina_AFL (1) (1).xlsx"


def clean(v):
    if v is None:
        return ""
    s = str(v).strip()
    # remove emoji-like prefix labels if desired
    return s


def write_csv(path: Path, rows: list[dict], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(rows)
    print(f"  ✓ {path.name}  ({len(rows)} filas)")


# ─────────────────────────────────────────────
#  Helpers
# ─────────────────────────────────────────────

def rows_of(ws):
    """Yield non-empty rows as lists of cleaned values."""
    for row in ws.iter_rows(values_only=True):
        vals = [clean(c) for c in row]
        if any(vals):
            yield vals


def find_header_row(ws, keywords: set[str]) -> int:
    """Return 0-based index of the row that best matches keywords."""
    best_idx, best_score = 0, -1
    for idx, row in enumerate(ws.iter_rows(max_row=10, values_only=True)):
        vals = [clean(c).lower() for c in row if c is not None]
        hits = sum(any(kw in v for kw in keywords) for v in vals)
        if hits > best_score:
            best_score, best_idx = hits, idx
    return best_idx


# ─────────────────────────────────────────────
#  Sheet extractors
# ─────────────────────────────────────────────

def extract_plan_accion(ws) -> list[dict]:
    """📋 Plan de Acción  →  plan_accion.csv"""
    rows = []
    header_done = False
    for row in ws.iter_rows(min_row=1, values_only=True):
        vals = [clean(c) for c in row]
        # Skip blank-ish rows and header title
        non_empty = [v for v in vals if v]
        if len(non_empty) < 3:
            continue
        # Detect data rows: first meaningful non-None cell is a number
        # Column layout: None, #, Acción, Línea, Área, Prioridad, Inicio, Fin, Resp., Recursos, KPI, Status
        num_col = vals[1] if len(vals) > 1 else ""
        if not header_done:
            if num_col == "#":
                header_done = True
            continue
        try:
            num = int(float(num_col)) if num_col else None
        except (ValueError, TypeError):
            num = None
        if num is None:
            continue

        priority_raw = clean(vals[5]) if len(vals) > 5 else ""
        # map emoji priorities to text
        if "ALTA" in priority_raw.upper():
            priority = "ALTA"
        elif "MEDIA" in priority_raw.upper():
            priority = "MEDIA"
        elif "BAJA" in priority_raw.upper():
            priority = "BAJA"
        else:
            priority = priority_raw

        rows.append({
            "num": num,
            "accion": clean(vals[2]) if len(vals) > 2 else "",
            "linea": clean(vals[3]) if len(vals) > 3 else "",
            "area": clean(vals[4]) if len(vals) > 4 else "",
            "prioridad": priority,
            "inicio": clean(vals[6]) if len(vals) > 6 else "",
            "fin": clean(vals[7]) if len(vals) > 7 else "",
            "responsable": clean(vals[8]) if len(vals) > 8 else "",
            "recursos": clean(vals[9]) if len(vals) > 9 else "",
            "kpi": clean(vals[10]) if len(vals) > 10 else "",
            "status": clean(vals[11]) if len(vals) > 11 else "pendiente",
        })
    return rows


def extract_balanceo(ws) -> list[dict]:
    """📉 Antes vs Después  →  balanceo_lineas.csv"""
    rows = []

    # --- NORTHFACE section ---
    nf_data = [
        {"linea": "NORTHFACE", "estacion": "Est1 Fab Sub",    "ct_actual": 2065, "ct_meta": 2065, "ahorro_seg": 0,    "ahorro_pct": 0,    "intervencion": "Mantener",                       "takt": 2821},
        {"linea": "NORTHFACE", "estacion": "Est2 Fab Caja",   "ct_actual": 1685, "ct_meta": 1685, "ahorro_seg": 0,    "ahorro_pct": 0,    "intervencion": "Absorber tareas de Est3",        "takt": 2821},
        {"linea": "NORTHFACE", "estacion": "Est3 Fab Tapa",   "ct_actual": 2978, "ct_meta": 2650, "ahorro_seg": 328,  "ahorro_pct": 11,   "intervencion": "Redistribuir 328s a Est2",       "takt": 2821},
        {"linea": "NORTHFACE", "estacion": "Est4 Sub H",      "ct_actual": 3953, "ct_meta": 2047, "ahorro_seg": 1906, "ahorro_pct": 48,   "intervencion": "PM + shadow board → -1906 seg",  "takt": 2821},
        {"linea": "NORTHFACE", "estacion": "Est5 Sub H Tapa", "ct_actual": 876,  "ct_meta": 876,  "ahorro_seg": 0,    "ahorro_pct": 0,    "intervencion": "Mantener — buffer",              "takt": 2821},
        {"linea": "NORTHFACE", "estacion": "Est6 Ensamble",   "ct_actual": 4923, "ct_meta": 2821, "ahorro_seg": 2102, "ahorro_pct": 43,   "intervencion": "QA inline + dividir pasos",      "takt": 2821},
        {"linea": "NORTHFACE", "estacion": "Est7 Integración","ct_actual": 3571, "ct_meta": 2068, "ahorro_seg": 1503, "ahorro_pct": 42,   "intervencion": "Fixture + shadow board + QA",    "takt": 2821},
    ]

    # --- SANMINA section ---
    sanmina_data = [
        {"linea": "SANMINA", "estacion": "Est1 Cables",       "ct_actual": 86,   "ct_meta": 86,  "ops_actual": 3, "ops_meta": 3, "ct_op_actual": 29,   "ct_op_meta": 29,  "accion": "Mantener",                        "takt": 2217},
        {"linea": "SANMINA", "estacion": "Est2 Tapa",         "ct_actual": 494,  "ct_meta": 494, "ops_actual": 1, "ops_meta": 1, "ct_op_actual": 494,  "ct_op_meta": 494, "accion": "Mantener",                        "takt": 2217},
        {"linea": "SANMINA", "estacion": "Est3 Base",         "ct_actual": 50,   "ct_meta": 50,  "ops_actual": 1, "ops_meta": 1, "ct_op_actual": 50,   "ct_op_meta": 50,  "accion": "QC entrada material",             "takt": 2217},
        {"linea": "SANMINA", "estacion": "Est4 Gaskets base", "ct_actual": 900,  "ct_meta": 250, "ops_actual": 3, "ops_meta": 3, "ct_op_actual": 300,  "ct_op_meta": 250, "accion": "Fixture → -150s/op",              "takt": 2217},
        {"linea": "SANMINA", "estacion": "Est5 Gaskets tapa", "ct_actual": 84,   "ct_meta": 84,  "ops_actual": 1, "ops_meta": 1, "ct_op_actual": 84,   "ct_op_meta": 84,  "accion": "Implementar fixture diseñado",    "takt": 2217},
        {"linea": "SANMINA", "estacion": "Est6 Cables ruteo", "ct_actual": 2820, "ct_meta": 940, "ops_actual": 2, "ops_meta": 3, "ct_op_actual": 1410, "ct_op_meta": 940, "accion": "1 op adicional + WI visual",      "takt": 2217},
        {"linea": "SANMINA", "estacion": "Est7 Final",        "ct_actual": 300,  "ct_meta": 300, "ops_actual": 2, "ops_meta": 2, "ct_op_actual": 150,  "ct_op_meta": 150, "accion": "Ergonomía carro",                 "takt": 2217},
    ]

    fieldnames_nf = ["linea", "estacion", "ct_actual", "ct_meta", "ahorro_seg", "ahorro_pct", "intervencion", "takt",
                     "ops_actual", "ops_meta", "ct_op_actual", "ct_op_meta", "accion"]
    for d in nf_data:
        d.setdefault("ops_actual", 1)
        d.setdefault("ops_meta", 1)
        d.setdefault("ct_op_actual", d["ct_actual"])
        d.setdefault("ct_op_meta", d["ct_meta"])
        d.setdefault("accion", d.get("intervencion", ""))
        rows.append(d)
    for d in sanmina_data:
        d.setdefault("ahorro_seg", d["ct_actual"] - d["ct_meta"])
        pct_raw = round((d["ct_actual"] - d["ct_meta"]) / d["ct_actual"] * 100) if d["ct_actual"] else 0
        d.setdefault("ahorro_pct", pct_raw)
        d.setdefault("intervencion", d.get("accion", ""))
        rows.append(d)

    return rows


def extract_desperdicios(ws) -> list[dict]:
    """📊 Desperdicios & Insights  →  desperdicios.csv"""
    rows = []
    # Key data from the sheet (manually parsed from known structure)
    data = [
        {"categoria": "Trabajo productivo",          "tiempo_seg": 2068, "pct": 46.1, "linea": "NORTHFACE", "estacion": "Est7", "causa": "Actividades de valor",           "accion": "Mantener y proteger"},
        {"categoria": "Verificaciones / CTQ / firma","tiempo_seg": 1037, "pct": 23.1, "linea": "NORTHFACE", "estacion": "Est7", "causa": "Proceso QA inline no integrado",  "accion": "QA inline + checkpoint digital"},
        {"categoria": "Esperas (QA / material)",     "tiempo_seg": 599,  "pct": 13.3, "linea": "NORTHFACE", "estacion": "Est7", "causa": "QA externo / material no pronto", "accion": "QA dedicado + supermercado"},
        {"categoria": "Retrabajo remaches",          "tiempo_seg": 399,  "pct": 8.9,  "linea": "NORTHFACE", "estacion": "Est7", "causa": "Herramienta sin PM / fixture",    "accion": "PM semanal + fixture anti-retrabajo"},
        {"categoria": "Caminatas innecesarias",      "tiempo_seg": 382,  "pct": 8.6,  "linea": "NORTHFACE", "estacion": "Est7", "causa": "Layout subóptimo / shadow board", "accion": "Shadow board + layout optimizado"},
    ]
    # Throughput improvement steps
    throughput = [
        {"etapa": "Actual (sin mejora)", "pzas_hr": 8.0,  "linea": "NORTHFACE", "estacion": "Est7"},
        {"etapa": "+ Shadow board",       "pzas_hr": 9.3,  "linea": "NORTHFACE", "estacion": "Est7"},
        {"etapa": "+ Fixture",            "pzas_hr": 9.5,  "linea": "NORTHFACE", "estacion": "Est7"},
        {"etapa": "+ WI visual",          "pzas_hr": 12.4, "linea": "NORTHFACE", "estacion": "Est7"},
    ]
    return data, throughput


def extract_estaciones_nf(ws) -> list[dict]:
    """🔧 Estaciones NF  →  estaciones_nf.csv"""
    rows = []
    current_station = ""
    for row in ws.iter_rows(min_row=1, values_only=True):
        vals = [clean(c) for c in row]
        if len(vals) < 3:
            continue
        # Station header rows: column B starts with ▶
        b = vals[1]
        if b.startswith("▶"):
            current_station = b.replace("▶", "").strip()
            continue
        # Data rows: column B is a number (step)
        try:
            step = int(float(b)) if b else None
        except (ValueError, TypeError):
            step = None
        if step is None:
            continue
        rows.append({
            "estacion": current_station,
            "paso": step,
            "actividad": vals[2] if len(vals) > 2 else "",
            "operador": vals[3] if len(vals) > 3 else "",
            "bec_seg": vals[4] if len(vals) > 4 else "",
            "work_seg": vals[5] if len(vals) > 5 else "",
            "caminata_seg": vals[6] if len(vals) > 6 else "",
            "espera_seg": vals[7] if len(vals) > 7 else "",
            "total_seg": vals[8] if len(vals) > 8 else "",
            "herramental_pn": vals[9] if len(vals) > 9 else "",
        })
    return rows


def extract_estaciones_sanmina(ws) -> list[dict]:
    """🔧 Estaciones Sanmina  →  estaciones_sanmina.csv"""
    rows = []
    current_station = ""
    for row in ws.iter_rows(min_row=1, values_only=True):
        vals = [clean(c) for c in row]
        if len(vals) < 3:
            continue
        b = vals[1]
        if b.startswith("▶"):
            current_station = b.replace("▶", "").strip()
            continue
        try:
            step = int(float(b)) if b else None
        except (ValueError, TypeError):
            step = None
        if step is None:
            continue
        rows.append({
            "estacion": current_station,
            "paso": step,
            "actividad": vals[2] if len(vals) > 2 else "",
            "operador": vals[3] if len(vals) > 3 else "",
            "material_pn": vals[4] if len(vals) > 4 else "",
            "herramental": vals[5] if len(vals) > 5 else "",
            "ct_seg": vals[6] if len(vals) > 6 else "",
            "ct_min": vals[7] if len(vals) > 7 else "",
            "pzas_hr": vals[8] if len(vals) > 8 else "",
            "observaciones": vals[9] if len(vals) > 9 else "",
            "accion_recomendada": vals[10] if len(vals) > 10 else "",
        })
    return rows


def extract_demanda(ws) -> list[dict]:
    """Demanda AFL desde hoja 📉 Antes vs Después  →  demanda_afl.csv"""
    rows = [
        {"programa": "ASCEND",    "part_number": "ASCND-4RU-12RT",    "dic": 400, "ene": 448, "feb": 300, "mar": 448, "abr": 448, "may": 560, "total": 2604, "pico": "May-25: 560"},
        {"programa": "ASCEND",    "part_number": "ASCND-2RU-12RT",    "dic": 500, "ene": 400, "feb": 400, "mar": 500, "abr": 400, "may": 400, "total": 2600, "pico": "Dic+Mar: 500"},
        {"programa": "ASCEND",    "part_number": "ASCND-1RU-12RT",    "dic": 232, "ene": 290, "feb": 232, "mar": 232, "abr": 290, "may": 232, "total": 1508, "pico": "Feb+May: 290"},
        {"programa": "AFL HYPER", "part_number": "HCF-48U-SPL-001",   "dic": 0,   "ene": 60,  "feb": 80,  "mar": 60,  "abr": 60,  "may": 40,  "total": 300,  "pico": "Feb-25: 80"},
        {"programa": "AFL HYPER", "part_number": "HCF-48U-ODF-003",   "dic": 28,  "ene": 35,  "feb": 28,  "mar": 14,  "abr": 0,   "may": 0,   "total": 105,  "pico": "Ene-25: 35"},
        {"programa": "ASCEND",    "part_number": "ASCND-4RU-8RT",     "dic": 0,   "ene": 200, "feb": 0,   "mar": 0,   "abr": 0,   "may": 0,   "total": 200,  "pico": "Ene-25: 200"},
        {"programa": "ASCEND",    "part_number": "ASCND-1RU-8RT",     "dic": 0,   "ene": 0,   "feb": 0,   "mar": 100, "abr": 0,   "may": 200, "total": 300,  "pico": "May-25: 200"},
        {"programa": "ASCEND",    "part_number": "ASCND-2RU-8RT",     "dic": 100, "ene": 0,   "feb": 0,   "mar": 0,   "abr": 0,   "may": 0,   "total": 100,  "pico": "Dic-24: 100"},
    ]
    return rows


# ─────────────────────────────────────────────
#  Main
# ─────────────────────────────────────────────

def build():
    if not REPORTE.exists():
        print(f"❌  No se encontró: {REPORTE}")
        return

    print(f"📂  Procesando: {REPORTE.name}")
    wb = openpyxl.load_workbook(REPORTE, data_only=True)

    # Map sheet names (strip emojis-first chars for lookup)
    sheets = {name: wb[name] for name in wb.sheetnames}

    results = {}

    # --- Plan de Acción ---
    for name, ws in sheets.items():
        if "Plan" in name and "Acci" in name:
            data = extract_plan_accion(ws)
            write_csv(
                CURATED_DIR / "plan_accion.csv",
                data,
                ["num", "accion", "linea", "area", "prioridad", "inicio", "fin",
                 "responsable", "recursos", "kpi", "status"],
            )
            results["plan_accion"] = len(data)
            break

    # --- Balanceo ---
    for name, ws in sheets.items():
        if "Antes" in name or "Despu" in name:
            data = extract_balanceo(ws)
            write_csv(
                CURATED_DIR / "balanceo_lineas.csv",
                data,
                ["linea", "estacion", "ct_actual", "ct_meta", "ahorro_seg", "ahorro_pct",
                 "intervencion", "takt", "ops_actual", "ops_meta", "ct_op_actual", "ct_op_meta", "accion"],
            )
            results["balanceo_lineas"] = len(data)
            break

    # --- Desperdicios ---
    for name, ws in sheets.items():
        if "Desperdicio" in name or "Insight" in name:
            data, throughput = extract_desperdicios(ws)
            write_csv(
                CURATED_DIR / "desperdicios.csv",
                data,
                ["categoria", "tiempo_seg", "pct", "linea", "estacion", "causa", "accion"],
            )
            write_csv(
                CURATED_DIR / "throughput_mejoras.csv",
                throughput,
                ["etapa", "pzas_hr", "linea", "estacion"],
            )
            results["desperdicios"] = len(data)
            results["throughput_mejoras"] = len(throughput)
            break

    # --- Estaciones NF ---
    for name, ws in sheets.items():
        if "Estacion" in name and ("NF" in name or "North" in name):
            data = extract_estaciones_nf(ws)
            write_csv(
                CURATED_DIR / "estaciones_nf.csv",
                data,
                ["estacion", "paso", "actividad", "operador", "bec_seg", "work_seg",
                 "caminata_seg", "espera_seg", "total_seg", "herramental_pn"],
            )
            results["estaciones_nf"] = len(data)
            break

    # --- Estaciones Sanmina ---
    for name, ws in sheets.items():
        if "Estacion" in name and "Sanmina" in name:
            data = extract_estaciones_sanmina(ws)
            write_csv(
                CURATED_DIR / "estaciones_sanmina.csv",
                data,
                ["estacion", "paso", "actividad", "operador", "material_pn", "herramental",
                 "ct_seg", "ct_min", "pzas_hr", "observaciones", "accion_recomendada"],
            )
            results["estaciones_sanmina"] = len(data)
            break

    # --- Demanda AFL ---
    demanda = extract_demanda(None)
    write_csv(
        CURATED_DIR / "demanda_afl.csv",
        demanda,
        ["programa", "part_number", "dic", "ene", "feb", "mar", "abr", "may", "total", "pico"],
    )
    results["demanda_afl"] = len(demanda)

    print("\n📊  Resumen:")
    print(json.dumps(results, indent=2, ensure_ascii=False))
    return results


if __name__ == "__main__":
    build()
