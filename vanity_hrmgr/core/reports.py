"""Generador de reportes Excel para exportación de datos."""

from openpyxl.styles import Font

from openpyxl import Workbook
from django.http import HttpResponse


def generate_excel_report(queryset, filename, headers, data_func):
    """Generate an Excel report from a queryset.

    Args:
        queryset: Django queryset to export
        filename: Name of the downloaded file
        headers: List of column headers
        data_func: Function that takes an object and returns a list of cell values
    """
    wb = Workbook()
    ws = wb.active
    ws.title = filename.replace(".xlsx", "")[:31]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    for row_num, obj in enumerate(queryset, 2):
        for col_num, value in enumerate(data_func(obj), 1):
            data = str(value) if value is not None else ""
            ws.cell(row=row_num, column=col_num, value=data)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f"attachment; filename={filename}"
    wb.save(response)
    return response
